from django.views.generic import CreateView, ListView, DetailView
from django.views import View
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponseForbidden
from django.contrib.auth.mixins import LoginRequiredMixin
from accounts.utils.rbac import has_any_role
from .models import YearlyInterestPool
from .forms import InterestPoolForm
from .services import distribute_interest
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime



class InterestPoolCreateView(CreateView):
    model = YearlyInterestPool
    form_class = InterestPoolForm
    template_name = "transact5_share_distrib/pool_created.html"
    success_url = reverse_lazy("transact5_share_distrib:pool-list")

    def form_valid(self, form):
        messages.success(self.request, "Interest pool created successfully.")
        return super().form_valid(form)

class OfficerRequiredMixin:
    def dispatch(self, request, *args, **kwargs):
        if not has_any_role(request.user, ["officer", "itadmin","manager"]):
            return HttpResponseForbidden("Not allowed")
        return super().dispatch(request, *args, **kwargs)


class InterestPoolListView(LoginRequiredMixin, ListView):
    model = YearlyInterestPool
    template_name = "transact5_share_distrib/pool_list.html"
    context_object_name = "pools"
    paginate_by = 10

    def get_queryset(self):
        return ( YearlyInterestPool.objects.select_related("source_account").order_by("-year"))


class InterestPoolDetailView(LoginRequiredMixin, DetailView):
    model = YearlyInterestPool
    template_name = "transact5_share_distrib/pool_detail.html"
    context_object_name = "pool"
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pool = self.object
        context["shares"] = ( pool.shares.select_related( "account", "account__member"  ) )
        context["can_distribute"] = pool.status == "approved"
        source_balance = pool.source_account.balance if pool.source_account else 0
        context["distributable_amount"] = min(pool.total_interest, source_balance)

        return context

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status == "distributed" and request.method != "GET":
            return redirect("transact5_share_distrib:pool-detail", pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)



class ApproveInterestPoolView(View):

    def post(self, request, pk):

        if not has_any_role(request.user, ["manager","itadmin"]):
            return HttpResponseForbidden("Managers only")

        pool = get_object_or_404(YearlyInterestPool, pk=pk)

        if pool.status != YearlyInterestPool.Status.PENDING:
            messages.warning(request, "Already processed.")
            return redirect("transact5_share_distrib:pool-detail", pk=pk)

        pool.status = YearlyInterestPool.Status.APPROVED
        pool.save(update_fields=["status"])

        messages.success(request, "Pool approved successfully.")
        return redirect("transact5_share_distrib:pool-detail", pk=pk)


class DistributeInterestView(LoginRequiredMixin, View):

    def post(self, request, pk):

        if not has_any_role(request.user, ["officer","itadmin", "manager"]):
            return HttpResponseForbidden("Not allowed")

        try:
            pool = get_object_or_404(YearlyInterestPool, pk=pk)

            distribute_interest(       year=pool.year,
                performed_by=request.user       )

            messages.success(request, "Interest distributed successfully.")

        except Exception as e:
            messages.error(request, str(e))

        return redirect("transact5_share_distrib:pool-detail", pk=pk)



class InterestPoolExportExcelView(LoginRequiredMixin, View):

    def get(self, request, pk):

        pool = get_object_or_404(YearlyInterestPool, pk=pk)
        shares = pool.shares.select_related("account", "account__member")

        wb = Workbook()
        ws = wb.active
        ws.title = f"Pool {pool.year}"

        # 🕒 Export timestamp
        export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([f"Exported At: {export_time}"])
        ws.append([])

        # 📌 Header row
        headers = [
            "Account Number",
            "Member",
            "Principal",
            "Ratio",
            "Interest Earned"
        ]
        ws.append(headers)

        # 🟡 Style headers (bold)
        bold_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")

        # 📊 Data rows
        start_row = 4

        for s in shares:
            ws.append([
                s.account.account_number,
                str(s.account.member),
                round(float(s.principal_snapshot), 2),
                round(float(s.ratio), 4),
                round(float(s.interest_earned), 2),
            ])

        # 📌 Summary section
        ws.append([])
        summary_row = ws.max_row + 1

        ws.append(["TOTAL INTEREST", "", "", "", round(float(pool.total_interest), 2)])
        ws.append(["TOTAL DISTRIBUTED", "", "", "", round(float(pool.distributed_amount), 2)])

        # 💰 Bold summary labels
        ws[f"A{summary_row}"].font = bold_font
        ws[f"A{summary_row+1}"].font = bold_font

        # 📏 Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 5

        # 📤 Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        filename = f"interest_pool_{pool.year}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response
##from transact5_share_distrib.models import YearlyInterestPool

#pool = YearlyInterestPool.objects.last()
#pool.year, pool.status, pool.total_interest, pool.distributed_amount
##from transact5_share_distrib.services import distribute_interest

##result = distribute_interest(year=pool.year)